"""
Options Volume Signal — Main Entry Point

Fetches options data for S&P 500 IT + Nasdaq-100 stocks, identifies the top-3
highest-volume contracts per expiry horizon (7d/30d/45d/60d/90d/180d/1y), computes
buy/sell signals and forecasted % moves, and exports a color-coded Excel file.

Usage:
    python main.py
    python main.py --tickers AAPL MSFT NVDA   # test with a subset
    python main.py --output my_report.xlsx
"""

import argparse
import json
import sys
from datetime import datetime, date
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

from universe import get_universe
from fetcher import fetch_all
from analyzer import analyze_all, HORIZONS, TOP_N
from exporter_html import write_html

# ---------------------------------------------------------------------------
# Color palette
# ---------------------------------------------------------------------------
GREEN_FILL   = PatternFill("solid", fgColor="C6EFCE")
RED_FILL     = PatternFill("solid", fgColor="FFC7CE")
YELLOW_FILL  = PatternFill("solid", fgColor="FFEB9C")
ORANGE_FILL  = PatternFill("solid", fgColor="FCE4D6")
BLUE_FILL    = PatternFill("solid", fgColor="BDD7EE")
GREY_FILL    = PatternFill("solid", fgColor="D9D9D9")
HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
SUBHDR_FILL  = PatternFill("solid", fgColor="2E75B6")
RANK_FILLS   = [
    PatternFill("solid", fgColor="DDEBF7"),
    PatternFill("solid", fgColor="EBF3FB"),
    PatternFill("solid", fgColor="F5F9FE"),
]

WHITE_FONT   = Font(color="FFFFFF", bold=True)
BOLD_FONT    = Font(bold=True)
CENTER       = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT         = Alignment(horizontal="left", vertical="center")

THIN = Side(style="thin", color="BFBFBF")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

HORIZONS_ORDER = list(HORIZONS.keys())  # ["7d","30d","45d","60d","90d","180d","1y"]


# ---------------------------------------------------------------------------
# Build flat rows for Detail sheet
# ---------------------------------------------------------------------------
def build_detail_rows(analyzed: list[dict]) -> list[dict]:
    rows = []
    for item in analyzed:
        ticker        = item["ticker"]
        price         = item["price"]
        earnings_date = item.get("earnings_date")
        for horizon in HORIZONS_ORDER:
            h_data           = item["horizons"].get(horizon, {})
            expiry           = h_data.get("expiry")
            earnings_in_win  = h_data.get("earnings_in_window", False)
            contracts        = h_data.get("contracts", [])
            for rank, contract in enumerate(contracts, 1):
                rows.append({
                    "Ticker":            ticker,
                    "Current Price":     price,
                    "Earnings Date":     earnings_date or "N/A",
                    "Horizon":           horizon,
                    "Rank":              rank,
                    "Expiry Date":       expiry or "N/A",
                    "⚠️ Earnings":       "YES" if earnings_in_win else "",
                    "Strike":            contract.get("strike"),
                    "Type":              contract.get("type"),
                    "Volume":            contract.get("volume"),
                    "Signal":            contract.get("signal"),
                    "Forecast %":        contract.get("forecast_pct"),
                    "Prev Strike":       contract.get("prev_strike"),
                    "Strike Δ":          contract.get("strike_delta"),
                    "Prev Signal":       contract.get("prev_signal"),
                    "Signal Flipped":    "YES" if contract.get("signal_flipped") else "",
                })
    return rows


# ---------------------------------------------------------------------------
# Write Detail sheet (long format)
# ---------------------------------------------------------------------------
def write_detail_sheet(ws, rows: list[dict]):
    headers = [
        "Ticker", "Current Price", "Earnings Date", "Horizon", "Rank",
        "Expiry Date", "⚠️ Earnings", "Strike", "Type", "Volume",
        "Signal", "Forecast %", "Prev Strike", "Strike Δ", "Prev Signal", "Signal Flipped",
    ]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    ws.row_dimensions[1].height = 30

    ORANGE_FONT = Font(color="833C00", bold=True)

    for r, row in enumerate(rows, 2):
        values = [
            row["Ticker"], row["Current Price"], row["Earnings Date"],
            row["Horizon"], row["Rank"], row["Expiry Date"], row["⚠️ Earnings"],
            row["Strike"], row["Type"], row["Volume"], row["Signal"],
            row["Forecast %"], row["Prev Strike"], row["Strike Δ"],
            row["Prev Signal"], row["Signal Flipped"],
        ]
        for c, val in enumerate(values, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = CENTER
            cell.border = THIN_BORDER

            col_name = headers[c - 1]
            if col_name in ("Signal", "Prev Signal"):
                if val == "BUY":
                    cell.fill = GREEN_FILL
                    cell.font = Font(color="375623", bold=True)
                elif val == "SELL":
                    cell.fill = RED_FILL
                    cell.font = Font(color="9C0006", bold=True)
                elif val == "HEDGE":
                    cell.fill = ORANGE_FILL
                    cell.font = ORANGE_FONT
            elif col_name == "Forecast %":
                if val is not None:
                    cell.number_format = '0.00"%"'
                    if val > 0:
                        cell.fill = GREEN_FILL
                    elif val < 0:
                        cell.fill = RED_FILL
                    else:
                        cell.fill = YELLOW_FILL
            elif col_name in ("Current Price", "Strike", "Prev Strike") and val is not None:
                cell.number_format = '"$"#,##0.00'
            elif col_name == "Strike Δ" and val is not None:
                cell.number_format = '"$"#,##0.00'
                if val > 0:
                    cell.fill = GREEN_FILL
                    cell.font = Font(color="375623")
                elif val < 0:
                    cell.fill = RED_FILL
                    cell.font = Font(color="9C0006")
            elif col_name == "Volume" and val is not None:
                cell.number_format = '#,##0'
            elif col_name == "Rank":
                fills = [GREEN_FILL, YELLOW_FILL, PatternFill("solid", fgColor="FCE4D6")]
                cell.fill = fills[int(val) - 1] if val and 1 <= int(val) <= 3 else GREY_FILL
            elif col_name == "⚠️ Earnings" and val == "YES":
                cell.fill = YELLOW_FILL
                cell.font = Font(color="7D5700", bold=True)
            elif col_name == "Signal Flipped" and val == "YES":
                cell.fill = PatternFill("solid", fgColor="E2EFDA")
                cell.font = Font(color="375623", bold=True)

    widths = [10, 13, 13, 9, 6, 12, 10, 10, 8, 10, 8, 11, 12, 10, 12, 14]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


# ---------------------------------------------------------------------------
# Write Summary sheet (wide format)
# ---------------------------------------------------------------------------
def write_summary_sheet(ws, analyzed: list[dict]):
    """
    Layout:
    Row 1: main headers (Ticker, Current Price, then horizon group headers spanning 3*(TOP_N) cols each)
    Row 2: sub-headers (for each horizon: Vol#1 Strike, Vol#1 Type, ... Vol#3 Forecast%)
    Row 3+: data
    """
    # Build column structure
    # Fixed cols: Ticker (1), Current Price (2)
    # Per horizon: TOP_N ranks × 5 fields = 15 cols per horizon
    FIELDS = ["Strike", "Type", "Volume", "Signal", "Forecast %"]
    N_FIELDS = len(FIELDS)

    fixed_cols = 2
    cols_per_horizon = TOP_N * N_FIELDS

    # Row 1: top-level headers
    ws.cell(row=1, column=1, value="Ticker").fill = HEADER_FILL
    ws.cell(row=1, column=1).font = WHITE_FONT
    ws.cell(row=1, column=1).alignment = CENTER
    ws.cell(row=1, column=1).border = THIN_BORDER

    ws.cell(row=1, column=2, value="Current Price").fill = HEADER_FILL
    ws.cell(row=1, column=2).font = WHITE_FONT
    ws.cell(row=1, column=2).alignment = CENTER
    ws.cell(row=1, column=2).border = THIN_BORDER

    # Merge Row 1 cells: Ticker spans rows 1-2, Current Price spans rows 1-2
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)

    horizon_label_map = {
        "7d":   "7 Days",
        "30d":  "1 Month",
        "45d":  "45 Days",
        "60d":  "60 Days",
        "90d":  "90 Days",
        "180d": "6 Months",
        "1y":   "1 Year",
    }

    for h_idx, horizon in enumerate(HORIZONS_ORDER):
        start_col = fixed_cols + 1 + h_idx * cols_per_horizon
        end_col   = start_col + cols_per_horizon - 1

        # Row 1: horizon group label
        cell = ws.cell(row=1, column=start_col,
                       value=f"{horizon_label_map[horizon]}  ({horizon})")
        cell.fill = SUBHDR_FILL
        cell.font = WHITE_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        ws.merge_cells(start_row=1, start_column=start_col,
                       end_row=1, end_column=end_col)

        # Row 2: rank sub-headers
        for rank in range(1, TOP_N + 1):
            for f_idx, field in enumerate(FIELDS):
                col = start_col + (rank - 1) * N_FIELDS + f_idx
                label = f"#{rank} {field}"
                cell = ws.cell(row=2, column=col, value=label)
                cell.fill = RANK_FILLS[rank - 1]
                cell.font = BOLD_FONT
                cell.alignment = CENTER
                cell.border = THIN_BORDER

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 36

    # Data rows
    for r_idx, item in enumerate(analyzed):
        row = r_idx + 3
        ticker = item["ticker"]
        price  = item["price"]

        cell = ws.cell(row=row, column=1, value=ticker)
        cell.font = BOLD_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER

        price_cell = ws.cell(row=row, column=2, value=price)
        price_cell.alignment = CENTER
        price_cell.border = THIN_BORDER
        if price is not None:
            price_cell.number_format = '"$"#,##0.00'

        for h_idx, horizon in enumerate(HORIZONS_ORDER):
            h_data    = item["horizons"].get(horizon, {})
            contracts = h_data.get("contracts", [])
            start_col = fixed_cols + 1 + h_idx * cols_per_horizon

            for rank_idx, contract in enumerate(contracts):
                base_col = start_col + rank_idx * N_FIELDS
                strike   = contract.get("strike")
                opt_type = contract.get("type")
                volume   = contract.get("volume")
                signal   = contract.get("signal")
                forecast = contract.get("forecast_pct")
                rank_fill = RANK_FILLS[rank_idx]

                # Strike
                c = ws.cell(row=row, column=base_col, value=strike)
                c.fill = rank_fill; c.alignment = CENTER; c.border = THIN_BORDER
                if strike is not None:
                    c.number_format = '"$"#,##0.00'

                # Type
                c = ws.cell(row=row, column=base_col + 1, value=opt_type)
                c.fill = rank_fill; c.alignment = CENTER; c.border = THIN_BORDER

                # Volume
                c = ws.cell(row=row, column=base_col + 2, value=volume)
                c.fill = rank_fill; c.alignment = CENTER; c.border = THIN_BORDER
                if volume is not None:
                    c.number_format = '#,##0'

                # Signal
                c = ws.cell(row=row, column=base_col + 3, value=signal)
                c.alignment = CENTER; c.border = THIN_BORDER
                if signal == "BUY":
                    c.fill = GREEN_FILL
                    c.font = Font(color="375623", bold=True)
                elif signal == "SELL":
                    c.fill = RED_FILL
                    c.font = Font(color="9C0006", bold=True)
                elif signal == "HEDGE":
                    c.fill = ORANGE_FILL
                    c.font = Font(color="833C00", bold=True)
                else:
                    c.fill = GREY_FILL

                # Forecast %
                c = ws.cell(row=row, column=base_col + 4, value=forecast)
                c.alignment = CENTER; c.border = THIN_BORDER
                if forecast is not None:
                    c.number_format = '0.00"%"'
                    if forecast > 0:
                        c.fill = GREEN_FILL
                        c.font = Font(color="375623")
                    elif forecast < 0:
                        c.fill = RED_FILL
                        c.font = Font(color="9C0006")
                    else:
                        c.fill = YELLOW_FILL
                else:
                    c.fill = GREY_FILL

    # Column widths
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 14
    col_widths = [10, 7, 9, 8, 10]  # Strike, Type, Volume, Signal, Forecast%
    for h_idx in range(len(HORIZONS_ORDER)):
        for rank in range(TOP_N):
            for f_idx, w in enumerate(col_widths):
                col = fixed_cols + 1 + h_idx * cols_per_horizon + rank * N_FIELDS + f_idx
                ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "C3"


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def parse_args():
    parser = argparse.ArgumentParser(description="Options Volume Signal Report")
    parser.add_argument(
        "--tickers", nargs="+", default=None,
        help="Optional subset of tickers to test with (e.g. AAPL MSFT NVDA)"
    )
    parser.add_argument(
        "--output", default=None,
        help="Output Excel filename (default: options_signals_YYYYMMDD_HHMM.xlsx)"
    )
    return parser.parse_args()


def main():
    args = parse_args()

    tickers = args.tickers if args.tickers else get_universe()
    now = datetime.now()
    timestamp = now.strftime("%Y%m%d_%H%M")
    display_timestamp = now.strftime("%B %-d, %Y at %-I:%M %p")

    reports_dir = Path("reports")
    reports_dir.mkdir(exist_ok=True)
    output_path = Path(args.output) if args.output else reports_dir / f"options_signals_{timestamp}.xlsx"
    snapshot_path = reports_dir / "latest_analysis.json"

    print(f"Options Volume Signal Report")
    print(f"Date: {display_timestamp}")
    print(f"Tickers: {len(tickers)}")
    print(f"Output:  {output_path}\n")

    # Load previous run snapshot for delta computation
    MAX_DELTA_AGE_DAYS = 3  # suppress delta if snapshot is older than this

    prev_data = None
    snapshot_info = {"age_days": None, "generated": None, "suppressed": False}

    if snapshot_path.exists():
        try:
            with open(snapshot_path, "r") as f:
                snapshot = json.load(f)
            # Support both old format (plain list) and new format (dict with metadata)
            if isinstance(snapshot, list):
                prev_data = snapshot
                snap_generated = None
                snap_age_days = None
            else:
                prev_data = snapshot.get("tickers", [])
                snap_generated = snapshot.get("generated")
                snap_age_days = None
                if snap_generated:
                    try:
                        snap_date = datetime.fromisoformat(snap_generated).date()
                        snap_age_days = (date.today() - snap_date).days
                    except ValueError:
                        pass

            suppressed = snap_age_days is not None and snap_age_days > MAX_DELTA_AGE_DAYS
            snapshot_info = {
                "age_days": snap_age_days,
                "generated": snap_generated,
                "suppressed": suppressed,
                "max_age": MAX_DELTA_AGE_DAYS,
            }

            age_str = f"{snap_age_days}d old" if snap_age_days is not None else "unknown age"
            warn = " — SUPPRESSED (too old)" if suppressed else ""
            print(f"Loaded previous snapshot: {snapshot_path} ({len(prev_data)} tickers, {age_str}{warn})\n")
        except Exception as e:
            print(f"[WARN] Could not load previous snapshot: {e}\n")

    # Step 1: Fetch
    print("--- Fetching data ---")
    fetch_results = fetch_all(tickers)

    # Step 2: Analyze (pass previous run for delta; suppress if snapshot too old)
    print("\n--- Analyzing ---")
    analyzed = analyze_all(
        fetch_results,
        prev_data=prev_data,
        suppress_delta=snapshot_info["suppressed"],
    )

    # Step 3: Build detail rows
    detail_rows = build_detail_rows(analyzed)
    print(f"Detail rows: {len(detail_rows)}")

    # Step 4: Write Excel
    print(f"\n--- Writing Excel: {output_path} ---")
    wb = Workbook()

    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    write_summary_sheet(ws_summary, analyzed)

    # Detail sheet
    ws_detail = wb.create_sheet("Detail")
    write_detail_sheet(ws_detail, detail_rows)

    wb.save(output_path)
    print(f"Excel report: {output_path.resolve()}")

    html_path = output_path.with_suffix(".html")
    write_html(analyzed, html_path, display_timestamp, snapshot_info=snapshot_info)
    print(f"\nDone! Open the HTML file in your browser to explore interactively.")

    # Save snapshot for next run's delta computation (includes timestamp for age check)
    try:
        with open(snapshot_path, "w") as f:
            json.dump({"generated": now.isoformat(), "tickers": analyzed}, f, default=str)
        print(f"Snapshot saved: {snapshot_path}")
    except Exception as e:
        print(f"[WARN] Could not save snapshot: {e}")

    # Quick stats
    total_buy     = sum(1 for r in detail_rows if r["Signal"] == "BUY")
    total_sell    = sum(1 for r in detail_rows if r["Signal"] == "SELL")
    total_hedge   = sum(1 for r in detail_rows if r["Signal"] == "HEDGE")
    total_na      = sum(1 for r in detail_rows if r["Signal"] is None)
    total_flipped = sum(1 for r in detail_rows if r["Signal Flipped"] == "YES")
    earnings_hits = sum(1 for r in detail_rows if r["⚠️ Earnings"] == "YES")
    print(f"\nSignal breakdown across all tickers/horizons/ranks:")
    print(f"  BUY:            {total_buy}")
    print(f"  SELL:           {total_sell}")
    print(f"  HEDGE:          {total_hedge}  (ITM — likely institutional hedges)")
    print(f"  N/A:            {total_na}")
    print(f"  Signal flips:   {total_flipped}  (signal changed vs previous run)")
    print(f"  Earnings flags: {earnings_hits}  (earnings within horizon window)")


if __name__ == "__main__":
    main()
